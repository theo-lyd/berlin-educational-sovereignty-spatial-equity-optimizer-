"""Phase 4 — Raw ingestion into DuckDB.

This script ingests the two source Excel workbooks into DuckDB without
performing destructive cleaning. It preserves source fidelity by loading
values as they appear in the workbook (including text markers such as
"k. A." and German-formatted cost strings).

It also writes:
- a DuckDB ingestion log table
- a CSV ingestion log
- a markdown validation report

Usage:
    python src/ingest_raw_duckdb.py \
        --input-dir /path/to/excel/files \
        --db-path data/warehouse/berlin_education.duckdb \
        --output-dir reports

Expected input files:
- od-eckdaten-allg-2024.xlsx
- schulbaukarte-2025-.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import duckdb
import pandas as pd
from openpyxl import load_workbook


@dataclass
class IngestionResult:
    source_file: str
    sheet_name: str
    table_name: str
    expected_rows: int
    loaded_rows: int
    loaded_columns: int
    status: str
    error_message: str | None = None
    ingested_at_utc: str | None = None


def slugify(value: str) -> str:
    value = value.strip().lower()
    chars = []
    last_was_underscore = False
    for ch in value:
        if ch.isalnum():
            chars.append(ch)
            last_was_underscore = False
        else:
            if not last_was_underscore:
                chars.append("_")
                last_was_underscore = True
    slug = "".join(chars).strip("_")
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug


def normalize_df_for_raw_load(df: pd.DataFrame) -> pd.DataFrame:
    # Preserve values as raw as possible. Convert NaN/NaT to None, but do not otherwise clean.
    return df.where(pd.notna(df), None)


def read_sheet_raw(xlsx_path: Path, sheet_name: str) -> Tuple[pd.DataFrame, int]:
    """Read a sheet with minimal transformation and return (dataframe, expected_rows)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    expected_rows = max(ws.max_row - 1, 0)
    wb.close()

    df = pd.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        dtype=object,
        engine="openpyxl",
    )
    df = normalize_df_for_raw_load(df)
    return df, expected_rows


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def ingest_workbook(
    con: duckdb.DuckDBPyConnection,
    xlsx_path: Path,
    source_label: str,
    raw_schema: str = "raw",
) -> List[IngestionResult]:
    results: List[IngestionResult] = []

    workbook_slug = slugify(xlsx_path.stem)

    # DuckDB schema for raw tables.
    con.execute(f"CREATE SCHEMA IF NOT EXISTS {raw_schema}")

    xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
    for sheet_name in xls.sheet_names:
        ts = datetime.now(timezone.utc).isoformat()
        table_name = f"{raw_schema}.raw__{workbook_slug}__{slugify(sheet_name)}"
        try:
            df, expected_rows = read_sheet_raw(xlsx_path, sheet_name)
            loaded_rows = len(df)
            loaded_columns = len(df.columns)

            # Overwrite any prior table so the process is idempotent.
            con.execute(f"DROP TABLE IF EXISTS {table_name}")
            con.register("_raw_df", df)
            con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM _raw_df")
            con.unregister("_raw_df")

            results.append(
                IngestionResult(
                    source_file=source_label,
                    sheet_name=sheet_name,
                    table_name=table_name,
                    expected_rows=expected_rows,
                    loaded_rows=loaded_rows,
                    loaded_columns=loaded_columns,
                    status="success",
                    ingested_at_utc=ts,
                )
            )
        except Exception as exc:  # noqa: BLE001
            results.append(
                IngestionResult(
                    source_file=source_label,
                    sheet_name=sheet_name,
                    table_name=table_name,
                    expected_rows=0,
                    loaded_rows=0,
                    loaded_columns=0,
                    status="failed",
                    error_message=str(exc),
                    ingested_at_utc=ts,
                )
            )

    return results


def write_duckdb_log_table(
    con: duckdb.DuckDBPyConnection,
    results: Iterable[IngestionResult],
) -> None:
    rows = [asdict(r) for r in results]
    log_df = pd.DataFrame(rows)
    con.register("_ingestion_log_df", log_df)
    con.execute("DROP TABLE IF EXISTS raw.raw_ingestion_log")
    con.execute("CREATE TABLE raw.raw_ingestion_log AS SELECT * FROM _ingestion_log_df")
    con.unregister("_ingestion_log_df")


def write_csv_log(results: Iterable[IngestionResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([asdict(r) for r in results]).to_csv(output_path, index=False)


def write_validation_report(results: Iterable[IngestionResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    results_list = list(results)

    total_sheets = len(results_list)
    success_count = sum(1 for r in results_list if r.status == "success")
    failed_count = total_sheets - success_count
    row_mismatch_count = sum(
        1
        for r in results_list
        if r.status == "success" and r.expected_rows != r.loaded_rows
    )

    lines = [
        "# Raw Load Validation Report",
        "",
        f"Generated at (UTC): {datetime.now(timezone.utc).isoformat()}",
        "",
        "## Summary",
        f"- Sheets processed: {total_sheets}",
        f"- Successful loads: {success_count}",
        f"- Failed loads: {failed_count}",
        f"- Row-count mismatches: {row_mismatch_count}",
        "",
        "## Per-sheet results",
        "| Source File | Sheet | Table | Expected Rows | Loaded Rows | Columns | Status |",
        "|---|---|---:|---:|---:|---:|---|",
    ]

    for r in results_list:
        lines.append(
            f"| {r.source_file} | {r.sheet_name} | `{r.table_name}` | {r.expected_rows} | {r.loaded_rows} | {r.loaded_columns} | {r.status} |"
        )
        if r.error_message:
            lines.append(f"  - Error: {r.error_message}")

    lines.append("")
    lines.append("## Interpretation")
    lines.append(
        "The raw layer is considered valid when each workbook sheet is loaded without destructive cleaning and row counts match the source workbook expectation."
    )
    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_validation_json(results: Iterable[IngestionResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = [asdict(r) for r in results]
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest source Excel files into DuckDB raw tables.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("data/raw"),
        help="Directory containing the source Excel workbooks.",
    )
    parser.add_argument(
        "--db-path",
        type=Path,
        default=Path("data/warehouse/berlin_education.duckdb"),
        help="Path to the DuckDB database file.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("reports"),
        help="Directory for ingestion logs and validation reports.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_dir: Path = args.input_dir
    db_path: Path = args.db_path
    output_dir: Path = args.output_dir

    workbook_map: Dict[str, str] = {
        "od-eckdaten-allg-2024.xlsx": "od-eckdaten-allg-2024.xlsx",
        "schulbaukarte-2025-.xlsx": "schulbaukarte-2025-.xlsx",
    }

    missing = [fname for fname in workbook_map if not (input_dir / fname).exists()]
    if missing:
        print(
            "Missing input files in input directory:\n- " + "\n- ".join(missing),
            file=sys.stderr,
        )
        return 1

    db_path.parent.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect(str(db_path))
    all_results: List[IngestionResult] = []

    try:
        for file_name in workbook_map.values():
            xlsx_path = input_dir / file_name
            results = ingest_workbook(con, xlsx_path, source_label=file_name)
            all_results.extend(results)

        write_duckdb_log_table(con, all_results)
        con.close()

        write_csv_log(all_results, output_dir / "raw_ingestion_log.csv")
        write_validation_report(all_results, output_dir / "raw_load_validation_report.md")
        write_validation_json(all_results, output_dir / "raw_load_validation_report.json")

        # Console summary.
        print("Raw ingestion completed.")
        for r in all_results:
            print(
                f"- {r.source_file} | {r.sheet_name} | {r.status} | expected={r.expected_rows} loaded={r.loaded_rows}"
            )
        return 0
    except Exception as exc:  # noqa: BLE001
        try:
            con.close()
        except Exception:
            pass
        print(f"Raw ingestion failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
